"""Data dictionary upload: parse, preview the merge, apply it.

XLSX layout (three sheets):
  Tables:  TABLE | DESCRIPTION
  Columns: TABLE | COLUMN | DESCRIPTION | CODES | FORMAT | SHARE_SAMPLES
  Joins:   LEFT_TABLE | LEFT_COLUMN | RIGHT_TABLE | RIGHT_COLUMN | DESCRIPTION

CSV layout (one flat file, KIND in {table, column, join}):
  KIND,TABLE,COLUMN,DESCRIPTION,CODES,FORMAT,SHARE_SAMPLES,RIGHT_TABLE,RIGHT_COLUMN

CODES uses "A=Active; P=Paid out". FORMAT is CYMD for numeric legacy dates.
"""

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

from .store import CatalogStore


class DictionaryError(Exception):
    pass


@dataclass
class Dictionary:
    tables: dict[str, str] = field(default_factory=dict)
    columns: dict[tuple[str, str], dict] = field(default_factory=dict)
    joins: list[dict] = field(default_factory=list)


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _truthy(value) -> bool:
    return str(value).strip().lower() in {"1", "y", "yes", "true"}


def _add_column_entry(doc, table, column, description, codes, fmt, share):
    doc.columns[(table.upper(), column.upper())] = {
        "description": description,
        "codes": codes,
        "value_format": fmt.upper() if fmt else None,
        "share_samples": 1 if share else 0,
    }


def parse_dictionary(data: bytes, filename: str) -> Dictionary:
    """Parse an uploaded dictionary file into a Dictionary document."""
    name = filename.lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(data)
    if name.endswith(".csv"):
        return _parse_csv(data)
    raise DictionaryError("Dictionary must be a .xlsx or .csv file")


def _parse_xlsx(data: bytes) -> Dictionary:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise DictionaryError("Could not read the workbook; is it a valid .xlsx?")
    doc = Dictionary()
    sheets = {s.lower(): s for s in wb.sheetnames}

    def rows_of(sheet_key):
        ws = wb[sheets[sheet_key]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return [], []
        keys = [_clean(h).upper() if _clean(h) else "" for h in header]
        return keys, rows

    if "tables" in sheets:
        keys, rows = rows_of("tables")
        for row in rows:
            rec = dict(zip(keys, row))
            table = _clean(rec.get("TABLE"))
            if table:
                doc.tables[table.upper()] = _clean(rec.get("DESCRIPTION"))
    if "columns" in sheets:
        keys, rows = rows_of("columns")
        for row in rows:
            rec = dict(zip(keys, row))
            table, column = _clean(rec.get("TABLE")), _clean(rec.get("COLUMN"))
            if table and column:
                _add_column_entry(
                    doc, table, column,
                    _clean(rec.get("DESCRIPTION")), _clean(rec.get("CODES")),
                    _clean(rec.get("FORMAT")), _truthy(rec.get("SHARE_SAMPLES")),
                )
    if "joins" in sheets:
        keys, rows = rows_of("joins")
        for row in rows:
            rec = dict(zip(keys, row))
            fields = [
                _clean(rec.get(k))
                for k in ("LEFT_TABLE", "LEFT_COLUMN", "RIGHT_TABLE", "RIGHT_COLUMN")
            ]
            if all(fields):
                doc.joins.append(
                    {
                        "left_table": fields[0].upper(),
                        "left_column": fields[1].upper(),
                        "right_table": fields[2].upper(),
                        "right_column": fields[3].upper(),
                        "description": _clean(rec.get("DESCRIPTION")),
                    }
                )
    if not (doc.tables or doc.columns or doc.joins):
        raise DictionaryError(
            "No Tables, Columns, or Joins sheet with data was found in the workbook"
        )
    return doc


def _parse_csv(data: bytes) -> Dictionary:
    doc = Dictionary()
    reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
    if not reader.fieldnames:
        raise DictionaryError("The CSV file is empty")
    for raw in reader:
        rec = {(k or "").strip().upper(): v for k, v in raw.items()}
        kind = (_clean(rec.get("KIND")) or "").lower()
        table = _clean(rec.get("TABLE"))
        if kind == "table" and table:
            doc.tables[table.upper()] = _clean(rec.get("DESCRIPTION"))
        elif kind == "column" and table and _clean(rec.get("COLUMN")):
            _add_column_entry(
                doc, table, _clean(rec.get("COLUMN")),
                _clean(rec.get("DESCRIPTION")), _clean(rec.get("CODES")),
                _clean(rec.get("FORMAT")), _truthy(rec.get("SHARE_SAMPLES")),
            )
        elif kind == "join":
            fields = [
                _clean(rec.get(k))
                for k in ("TABLE", "COLUMN", "RIGHT_TABLE", "RIGHT_COLUMN")
            ]
            if all(fields):
                doc.joins.append(
                    {
                        "left_table": fields[0].upper(),
                        "left_column": fields[1].upper(),
                        "right_table": fields[2].upper(),
                        "right_column": fields[3].upper(),
                        "description": _clean(rec.get("DESCRIPTION")),
                    }
                )
    if not (doc.tables or doc.columns or doc.joins):
        raise DictionaryError("No table, column, or join rows were found in the CSV")
    return doc


def merge_preview(store: CatalogStore, connection_id: int, doc: Dictionary) -> dict:
    """Describe what applying the dictionary would change, without changing it.

    Dictionary entries for tables or columns that introspection does not
    know about are reported as unknown and will be skipped: structure comes
    from the live schema, meaning comes from the dictionary.
    """
    catalog = store.catalog_for_connection(connection_id)
    known_tables = {t["table_name"].upper(): t for t in catalog}
    known_columns = {
        (t["table_name"].upper(), c["column_name"].upper()): c
        for t in catalog
        for c in t["columns"]
    }
    added, overwritten, unknown = [], [], []

    def record(target, fld, old, new):
        if new is None or new == old:
            return
        bucket = overwritten if old else added
        bucket.append({"target": target, "field": fld, "old": old, "new": new})

    for table, description in doc.tables.items():
        if table not in known_tables:
            unknown.append(f"Table {table}")
            continue
        record(table, "description", known_tables[table]["description"], description)
    for (table, column), entry in doc.columns.items():
        existing = known_columns.get((table, column))
        if existing is None:
            unknown.append(f"Column {table}.{column}")
            continue
        target = f"{table}.{column}"
        record(target, "description", existing["description"], entry["description"])
        record(target, "codes", existing["codes"], entry["codes"])
        record(target, "format", existing["value_format"], entry["value_format"])
        if entry["share_samples"] and not existing["share_samples"]:
            added.append(
                {"target": target, "field": "share samples", "old": None, "new": "on"}
            )
    valid_joins = 0
    for j in doc.joins:
        if (j["left_table"], j["left_column"]) not in known_columns:
            unknown.append(f"Join column {j['left_table']}.{j['left_column']}")
        elif (j["right_table"], j["right_column"]) not in known_columns:
            unknown.append(f"Join column {j['right_table']}.{j['right_column']}")
        else:
            valid_joins += 1
    return {
        "added": added,
        "overwritten": overwritten,
        "unknown": unknown,
        "join_count": valid_joins,
    }


def apply_dictionary(store: CatalogStore, connection_id: int, doc: Dictionary) -> dict:
    """Apply the dictionary; returns the preview of what was changed."""
    preview = merge_preview(store, connection_id, doc)
    catalog = store.catalog_for_connection(connection_id)
    known_tables = {t["table_name"].upper(): t for t in catalog}
    known_columns = {
        (t["table_name"].upper(), c["column_name"].upper()): c
        for t in catalog
        for c in t["columns"]
    }
    for table, description in doc.tables.items():
        if table in known_tables and description:
            store.update_table(known_tables[table]["id"], description)
    for (table, column), entry in doc.columns.items():
        existing = known_columns.get((table, column))
        if existing is None:
            continue
        fields = {
            k: v
            for k, v in (
                ("description", entry["description"]),
                ("codes", entry["codes"]),
                ("value_format", entry["value_format"]),
            )
            if v is not None
        }
        if entry["share_samples"]:
            fields["share_samples"] = 1
        store.update_column(existing["id"], fields)
    joins = [
        j
        for j in doc.joins
        if (j["left_table"], j["left_column"]) in known_columns
        and (j["right_table"], j["right_column"]) in known_columns
    ]
    if joins:
        store.replace_join_paths(connection_id, joins)
    return preview
