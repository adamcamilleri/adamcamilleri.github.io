"""Generate demo_dictionary.xlsx: the business dictionary for MIPROD.

This is the file a real customer would build themselves; shipping it means
the demo works instantly while the upload flow stays exercisable.
"""

from openpyxl import Workbook

TABLES = [
    ("LNMAST", "Loan master. One row per mortgage loan."),
    ("BRWR", "Borrowers on a loan. Join to LNMAST on LNNBR."),
    ("PRPTY", "Property securing the loan. One row per loan."),
    ("PYHIST", "Payment history. One row per payment event."),
    ("RNWL", "Renewal offers and their outcomes."),
    ("TXARR", "Property tax arrears records."),
    ("STSCDS", "Generic code lookup table. One row per code value."),
]

COLUMNS = [
    ("LNMAST", "LNNBR", "Loan number", None, None, ""),
    ("LNMAST", "LNSTCD", "Loan status",
     "A=Active; P=Paid out; D=Default; W=Written off", None, ""),
    ("LNMAST", "LNBAL", "Current principal balance", None, None, ""),
    ("LNMAST", "ORGAMT", "Original principal amount", None, None, ""),
    ("LNMAST", "INTRT", "Annual interest rate, percent", None, None, ""),
    ("LNMAST", "ORGDTE", "Origination date", None, "CYMD", ""),
    ("LNMAST", "MTDTE", "Maturity date", None, "CYMD", ""),
    ("LNMAST", "PRDCD", "Product",
     "F1=1-year fixed; F3=3-year fixed; F5=5-year fixed; V5=5-year variable; HL=HELOC",
     None, ""),
    ("LNMAST", "INVCD", "Investor",
     "B1=Balance sheet; S1=Securitized pool 1; S2=Securitized pool 2", None, ""),
    ("LNMAST", "PYMTAMT", "Scheduled monthly payment", None, None, ""),
    ("BRWR", "LNNBR", "Loan number", None, None, ""),
    ("BRWR", "BRSEQ", "Borrower sequence on the loan", None, None, ""),
    ("BRWR", "BRTYP", "Borrower type", "P=Primary; C=Co-borrower", None, ""),
    ("BRWR", "BRFNM", "Borrower first name", None, None, ""),
    ("BRWR", "BRLNM", "Borrower last name", None, None, ""),
    ("PRPTY", "LNNBR", "Loan number", None, None, ""),
    ("PRPTY", "PRVCD", "Province",
     "ON=Ontario; QC=Quebec; BC=British Columbia; AB=Alberta; MB=Manitoba; "
     "SK=Saskatchewan; NS=Nova Scotia; NB=New Brunswick; NL=Newfoundland and Labrador",
     None, "Y"),
    ("PRPTY", "CTYNM", "City name", None, None, ""),
    ("PRPTY", "PSTLCD", "Postal code", None, None, ""),
    ("PRPTY", "PRPTYP", "Property type",
     "D=Detached; S=Semi-detached; T=Townhouse; C=Condo", None, ""),
    ("PYHIST", "LNNBR", "Loan number", None, None, ""),
    ("PYHIST", "PYDTE", "Payment date", None, "CYMD", ""),
    ("PYHIST", "PYAMT", "Payment amount; negative means a reversal", None, None, ""),
    ("PYHIST", "PYTYP", "Payment type",
     "R=Regular; P=Privilege prepayment; N=NSF reversal", None, ""),
    ("RNWL", "LNNBR", "Loan number", None, None, ""),
    ("RNWL", "RNDTE", "Renewal offer date", None, "CYMD", ""),
    ("RNWL", "OFRRT", "Offered renewal rate, percent", None, None, ""),
    ("RNWL", "RNSTCD", "Renewal status",
     "O=Offered; A=Accepted; D=Declined; E=Expired", None, ""),
    ("TXARR", "LNNBR", "Loan number", None, None, ""),
    ("TXARR", "TAXYR", "Tax year", None, None, ""),
    ("TXARR", "ARRAMT", "Arrears amount", None, None, ""),
    ("TXARR", "ARRDTE", "Date the arrears were recorded", None, "CYMD", ""),
    ("TXARR", "ARRSTS", "Arrears status", "O=Outstanding; P=Paid", None, ""),
    ("STSCDS", "CDTYP", "Code type; matches the coded column name", None, None, ""),
    ("STSCDS", "CDVAL", "Code value", None, None, ""),
    ("STSCDS", "CDDSC", "Code description", None, None, ""),
]

JOINS = [
    ("LNMAST", "LNNBR", "BRWR", "LNNBR", "Borrowers on the loan"),
    ("LNMAST", "LNNBR", "PRPTY", "LNNBR", "Property securing the loan"),
    ("LNMAST", "LNNBR", "PYHIST", "LNNBR", "Payments on the loan"),
    ("LNMAST", "LNNBR", "RNWL", "LNNBR", "Renewal offers for the loan"),
    ("LNMAST", "LNNBR", "TXARR", "LNNBR", "Tax arrears on the loan"),
]


def write_dictionary(path) -> None:
    wb = Workbook()
    tables = wb.active
    tables.title = "Tables"
    tables.append(["TABLE", "DESCRIPTION"])
    for row in TABLES:
        tables.append(list(row))

    columns = wb.create_sheet("Columns")
    columns.append(["TABLE", "COLUMN", "DESCRIPTION", "CODES", "FORMAT", "SHARE_SAMPLES"])
    for row in COLUMNS:
        columns.append(list(row))

    joins = wb.create_sheet("Joins")
    joins.append(["LEFT_TABLE", "LEFT_COLUMN", "RIGHT_TABLE", "RIGHT_COLUMN", "DESCRIPTION"])
    for row in JOINS:
        joins.append(list(row))

    wb.save(path)
