"""XLSX and CSV export of query results.

XLSX gets formatted headers ("Business name (PHYS)"), a frozen top row,
CYMD columns written as real dates, and a decoded label column after each
coded column.
"""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..connectors.base import QueryResult
from ..cymd import cymd_to_date

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")


def _header_label(col: dict) -> str:
    if col["business_name"]:
        return f"{col['business_name']} ({col['name']})"
    return col["name"]


def export_xlsx(result: QueryResult, columns: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = []
    for col in columns:
        headers.append(_header_label(col))
        if col["codes"]:
            headers.append(f"{col['name']} decoded")
    for idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"

    for row in result.rows:
        out_row = []
        for value, col in zip(row, columns):
            if col["value_format"] == "CYMD":
                as_date = cymd_to_date(value)
                out_row.append(as_date if as_date else value)
            else:
                out_row.append(value)
            if col["codes"]:
                out_row.append(col["codes"].get(str(value)) if value is not None else None)
        ws.append(out_row)

    for idx, col in enumerate(_expanded_columns(columns), start=1):
        letter = get_column_letter(idx)
        if col.get("value_format") == "CYMD":
            for cell in ws[letter][1:]:
                if cell.value is not None and not isinstance(cell.value, (int, float)):
                    cell.number_format = "yyyy-mm-dd"
        width = max(12, min(40, len(str(ws.cell(row=1, column=idx).value or "")) + 2))
        ws.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _expanded_columns(columns):
    for col in columns:
        yield col
        if col["codes"]:
            yield {"name": f"{col['name']} decoded", "codes": None, "value_format": None}


def export_csv(result: QueryResult, columns: list[dict]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    headers = []
    for col in columns:
        headers.append(_header_label(col))
        if col["codes"]:
            headers.append(f"{col['name']} decoded")
    writer.writerow(headers)
    for row in result.rows:
        out_row = []
        for value, col in zip(row, columns):
            if col["value_format"] == "CYMD":
                as_date = cymd_to_date(value)
                out_row.append(as_date.isoformat() if as_date else value)
            else:
                out_row.append(value)
            if col["codes"]:
                out_row.append(col["codes"].get(str(value), "") if value is not None else "")
        writer.writerow(out_row)
    return buffer.getvalue().encode("utf-8-sig")
