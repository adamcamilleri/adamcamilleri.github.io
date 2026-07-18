"""End to end: seed a small demo db, run a canned request through the
generation-mocked pipeline, execute it, and export a working XLSX."""

import io
import json

import pytest
from openpyxl import load_workbook

from app.catalog.store import CatalogStore
from app.connectors.sqlite_connector import SQLiteConnector
from app.delivery.decode import describe_columns
from app.delivery.export import export_xlsx
from app.pipeline.generate import generate_sql
from app.seed.make_dictionary import write_dictionary
from app.seed.seed_demo import bootstrap_app_db, seed_demo_db

CANNED_REPLY = json.dumps(
    {
        "sql": (
            "SELECT CAST(PYDTE / 100 AS INTEGER) AS pay_month, SUM(PYAMT) AS total_privilege "
            "FROM PYHIST WHERE PYTYP = 'P' AND PYDTE BETWEEN 20250101 AND 20251231 "
            "GROUP BY pay_month ORDER BY pay_month"
        ),
        "explanation": (
            "Sums privilege prepayments from the payment history for 2025, "
            "grouped by calendar month."
        ),
    }
)


@pytest.fixture(scope="module")
def demo(tmp_path_factory):
    root = tmp_path_factory.mktemp("querydesk")
    demo_db = root / "demo_mars.db"
    dictionary = root / "demo_dictionary.xlsx"
    app_db = root / "app.db"
    seed_demo_db(demo_db, loan_count=150, rng_seed=7)
    write_dictionary(dictionary)
    connection_id = bootstrap_app_db(app_db, demo_db, dictionary)
    from app.appdb import make_engine

    engine = make_engine(app_db)
    yield {
        "store": CatalogStore(engine),
        "connector": SQLiteConnector(str(demo_db)),
        "connection_id": connection_id,
    }
    engine.dispose()


def test_pipeline_generates_validates_and_runs(demo):
    outcome = generate_sql(
        demo["store"],
        demo["connection_id"],
        demo["connector"],
        "Total privilege payments by month for 2025",
        complete=lambda system, messages: CANNED_REPLY,
    )
    assert outcome["ok"], outcome
    assert "LIMIT 5000" in outcome["sql"]

    result = demo["connector"].execute_readonly(
        outcome["sql"], row_limit=5000, timeout_s=30
    )
    assert result.columns == ["pay_month", "total_privilege"]
    assert 0 < len(result.rows) <= 12
    for month, total in result.rows:
        assert 202501 <= month <= 202512
        assert total > 0


def test_repair_loop_recovers_from_bad_first_attempt(demo):
    replies = iter(
        [
            json.dumps({"sql": "SELECT NOPE FROM LNMAST", "explanation": "bad"}),
            CANNED_REPLY,
        ]
    )
    seen_repair_prompts = []

    def scripted(system, messages):
        if len(messages) > 1:
            seen_repair_prompts.append(messages[-1]["content"])
        return next(replies)

    outcome = generate_sql(
        demo["store"],
        demo["connection_id"],
        demo["connector"],
        "Total privilege payments by month for 2025",
        complete=scripted,
    )
    assert outcome["ok"]
    assert outcome["attempts"] == ["Unknown column: NOPE"]
    assert "NOPE" in seen_repair_prompts[0]


def test_xlsx_export_opens_with_correct_headers(demo):
    sql = (
        "SELECT LNNBR, LNSTCD, LNBAL, MTDTE FROM LNMAST "
        "WHERE LNSTCD = 'A' ORDER BY LNNBR LIMIT 20"
    )
    result = demo["connector"].execute_readonly(sql, row_limit=5000, timeout_s=30)
    columns = describe_columns(demo["store"], demo["connection_id"], result.columns)
    payload = export_xlsx(result, columns)

    wb = load_workbook(io.BytesIO(payload))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == [
        "Loan number (LNNBR)",
        "Loan status (LNSTCD)",
        "LNSTCD decoded",
        "Current principal balance (LNBAL)",
        "Maturity date (MTDTE)",
    ]
    assert ws.freeze_panes == "A2"
    first = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert first[1] == "A"
    assert first[2] == "Active"
    # CYMD maturity dates land in Excel as real dates.
    maturity = ws.cell(row=2, column=5).value
    assert hasattr(maturity, "year")
    assert 2020 <= maturity.year <= 2032
